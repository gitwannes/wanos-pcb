# --- file: helpers/update_components_xlsx_s1.py ---
"""One-shot: align components.xlsx with S1 kickoff locks."""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook


def main() -> None:
    path = Path(__file__).resolve().parents[1] / "projects" / "wanos-board" / "components.xlsx"
    wb = load_workbook(path)
    ws = wb.active

    for i, row in enumerate(list(ws.iter_rows(min_row=2)), start=2):
        if row[0].value == "J15":
            ws.delete_rows(i, 1)
            break

    by_ref: dict[str, tuple] = {}
    for row in ws.iter_rows(min_row=2):
        ref = row[0].value
        if ref:
            by_ref[str(ref)] = row

    if "J1" in by_ref:
        by_ref["J1"][5].value = "C6990958"

    if "J41" in by_ref:
        by_ref["J41"][2].value = "DNP"
        by_ref["J41"][3].value = "DNP"
        by_ref["J41"][6].value = "DNP v1 - Pi fed via J40 header 5V"

    if "J17" not in by_ref:
        ws.append(
            [
                "J17",
                1,
                "KF301-2P",
                "TerminalBlock_Phoenix:TerminalBlock_MKDS-2-5.08_1x02_P5.08mm",
                "KF301-2P",
                "C474892",
                "Connector",
                "5V screw input",
                "5V",
                "TH",
                "TH",
            ]
        )

    extras = [
        ["F1", 1, "2A", "Fuse:Fuse_1206-3216Metric", "2A", "C72043", "Protection", "5V input fuse", "5V", "1206", "SMT"],
        ["F2", 1, "500mA", "Fuse:Fuse_1206-3216Metric", "500mA", "C369233", "Protection", "HDMI pin18 polyfuse", "5V", "1206", "SMT"],
        ["D1", 1, "BZX85C5V6", "Diode_THT:D_DO-41_SOD81_P7.62mm_Horizontal", "BZX85C5V6", "C81598", "Protection", "5V zener", "5V", "DO-41", "TH"],
        ["D2", 1, "1N4001", "Diode_THT:D_DO-41_SOD81_P7.62mm_Horizontal", "1N4001", "C81598", "Protection", "5V reverse block", "5V", "DO-41", "TH"],
        ["Q5", 1, "PN2222A", "Package_TO_SOT_SMD:SOT-23", "PN2222A-TA", "C2150", "Transistor", "SSR master safety BCM4", "12V", "SOT-23", "SMT"],
        ["R34", 1, "10k", "Resistor_SMD:R_0603", "0603WAF1002T5E", None, "Resistor", "EXP_B P3 pull-up", "3.3V", "0603", "SMT"],
        ["R35", 1, "10k", "Resistor_SMD:R_0603", "0603WAF1002T5E", None, "Resistor", "EXP_B P4 pull-up", "3.3V", "0603", "SMT"],
        ["R36", 1, "10k", "Resistor_SMD:R_0603", "0603WAF1002T5E", None, "Resistor", "EXP_B P5 pull-up", "3.3V", "0603", "SMT"],
    ]
    existing = {row[0].value for row in ws.iter_rows(min_row=2) if row[0].value}
    for item in extras:
        if item[0] not in existing:
            ws.append(item)

    by_ref = {str(row[0].value): row for row in ws.iter_rows(min_row=2) if row[0].value}
    if "R17-R28" in by_ref:
        by_ref["R17-R28"][2].value = "1k0"
    if "R29-R31" in by_ref:
        by_ref["R29-R31"][2].value = "R29/R31=2k0 R30=6k8"
    if "R32" in by_ref:
        by_ref["R32"][2].value = "1k5"

    wb.save(path)
    print(f"Updated {path}")


if __name__ == "__main__":
    main()
